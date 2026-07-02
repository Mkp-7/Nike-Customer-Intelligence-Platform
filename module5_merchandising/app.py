"""
Module 5 — Merchandising Line Plan
Uses Nike's own public API (api.nike.com) — completely free, no key needed.
Mirrors the work of a Nike Merchandising Information Analyst.
"""

import os, sys, json, urllib.request, urllib.parse
import pandas as pd
import plotly.express as px
import streamlit as st
from io import BytesIO

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE_DIR)

NIKE_API_BASE = "https://api.nike.com/cic/browse/v2"

CATEGORIES = {
    "Men's Running":    "010794e5-35fe-4e32-aaff-cd2c74f89d61",
    "Women's Running":  "16633190-45e5-4830-a068-232ac7aea82c",
    "Men's Basketball": "0f64ecc7-d624-4e91-b171-b83a03dd8550",
    "Jordan":           "5b6a9350-a3bb-4e8b-b660-f87e64f02700",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json",
    "Accept-Language": "en-US,en;q=0.9",
    "Origin": "https://www.nike.com",
    "Referer": "https://www.nike.com/",
}


def fetch_nike_category(filter_id: str, count: int = 24) -> list:
    endpoint = (f"/product_feed/rollup_threads/v2"
                f"?filter=marketplace(US)&filter=language(en)"
                f"&filter=employeePrice(true)&filter=attributeIds({filter_id})"
                f"&anchor=0&count={count}")
    params = urllib.parse.urlencode({
        "queryid": "products",
        "anonymousId": "7CC266B713D36CCC7275B33B6E4F9206",
        "country": "us",
        "endpoint": endpoint,
        "language": "en",
        "localizedRangeStr": "{lowestPrice} — {highestPrice}",
    })
    url = f"{NIKE_API_BASE}?{params}"
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read().decode("utf-8"))
        return data.get("data", {}).get("products", {}).get("products", [])
    except Exception as ex:
        st.warning(f"Could not fetch category: {ex}")
        return []


def parse_products(raw: list, category_label: str) -> list:
    rows = []
    for p in raw:
        if not isinstance(p, dict):
            continue
        price        = p.get("price") or {}
        current_p    = price.get("currentPrice") or price.get("fullPrice") or ""
        full_p       = price.get("fullPrice") or current_p
        in_stock     = p.get("inStock", True)
        label        = str(p.get("label", ""))
        on_sale      = False
        try:
            on_sale = float(str(current_p).replace(",","") or 0) < float(str(full_p).replace(",","") or 0)
        except Exception:
            pass
        if "New" in label:      status = "NEW"
        elif on_sale:           status = "SALE"
        elif not in_stock:      status = "OUT OF STOCK"
        else:                   status = "ACTIVE"
        gender = "Men" if "Men" in category_label else "Women" if "Women" in category_label else "Unisex"
        url = str(p.get("url", "")).replace("{countryLang}", "https://www.nike.com")
        rows.append({
            "SKU / Product ID":  p.get("pid", ""),
            "Product Name":      p.get("title", ""),
            "Subtitle":          p.get("subtitle", ""),
            "Colorway":          p.get("colorDescription", ""),
            "Gender":            gender,
            "Category":          category_label,
            "Retail Price ($)":  current_p,
            "Full Price ($)":    full_p,
            "Status":            status,
            "In Stock":          "✅" if in_stock else "❌",
            "URL":               url,
        })
    return rows


@st.cache_data(show_spinner=False, ttl=3600)
def load_line_plan() -> pd.DataFrame:
    all_rows = []
    for label, fid in CATEGORIES.items():
        raw  = fetch_nike_category(fid, count=24)
        rows = parse_products(raw, label)
        all_rows.extend(rows)
    if not all_rows:
        return pd.DataFrame()
    df = pd.DataFrame(all_rows)
    for col in ["Retail Price ($)", "Full Price ($)"]:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(",",""), errors="coerce")
    return df


def export_excel(df: pd.DataFrame) -> BytesIO:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output   = BytesIO()
    export_df = df.drop(columns=["URL"], errors="ignore")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Line Plan", index=False)
        ws = writer.sheets["Line Plan"]
        hf = PatternFill("solid", fgColor="111827")
        hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        bs = Side(style="thin", color="E5E7EB")
        tb = Border(left=bs, right=bs, top=bs, bottom=bs)
        sfills = {
            "NEW":          PatternFill("solid", fgColor="D1FAE5"),
            "SALE":         PatternFill("solid", fgColor="FEE2E2"),
            "ACTIVE":       PatternFill("solid", fgColor="EFF6FF"),
            "OUT OF STOCK": PatternFill("solid", fgColor="F3F4F6"),
        }
        for c in range(1, len(export_df.columns)+1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hf; cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = tb
        ws.row_dimensions[1].height = 28
        for ri, row in enumerate(export_df.itertuples(index=False), 2):
            sf = sfills.get(getattr(row,"Status","ACTIVE"), PatternFill("solid", fgColor="FFFFFF"))
            for ci in range(1, len(export_df.columns)+1):
                cell = ws.cell(row=ri, column=ci)
                cell.fill = sf; cell.font = Font(name="Arial", size=9)
                cell.border = tb; cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[ri].height = 18
        for i, w in enumerate([18,30,20,20,10,20,14,14,14,10], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(export_df.columns))}{len(export_df)+1}"

        if "Category" in df.columns and "Retail Price ($)" in df.columns:
            pp = df.groupby("Category")["Retail Price ($)"].agg(Products="count", Avg="mean", Min="min", Max="max").round(2).reset_index()
            pp.columns = ["Category","# Products","Avg Price ($)","Min Price ($)","Max Price ($)"]
            pp.to_excel(writer, sheet_name="Price Analysis", index=False)
            ws2 = writer.sheets["Price Analysis"]
            for c in range(1,6):
                ws2.cell(row=1,column=c).fill=hf; ws2.cell(row=1,column=c).font=hfont
                ws2.column_dimensions[get_column_letter(c)].width=16

        if "Status" in df.columns:
            sp = df.groupby(["Category","Status"]).size().reset_index(name="SKUs")
            sp.to_excel(writer, sheet_name="Status Summary", index=False)
            ws3 = writer.sheets["Status Summary"]
            for c in range(1,4):
                ws3.cell(row=1,column=c).fill=hf; ws3.cell(row=1,column=c).font=hfont
                ws3.column_dimensions[get_column_letter(c)].width=22

        if "Retail Price ($)" in df.columns:
            tmp = df.copy()
            tmp["Price Band"] = pd.cut(tmp["Retail Price ($)"],
                bins=[0,75,100,150,200,999], labels=["Under $75","$75–$100","$100–$150","$150–$200","$200+"])
            bp = tmp.groupby(["Category","Price Band"], observed=True).size().reset_index(name="SKUs")
            bp.to_excel(writer, sheet_name="Price Band", index=False)
            ws4 = writer.sheets["Price Band"]
            for c in range(1,4):
                ws4.cell(row=1,column=c).fill=hf; ws4.cell(row=1,column=c).font=hfont
                ws4.column_dimensions[get_column_letter(c)].width=22

    output.seek(0)
    return output


def show():
    st.markdown("## 📋 Merchandising Line Plan")
    st.markdown("Live Nike product catalog via Nike's public API — no API key required. "
                "SKU-level pricing, assortment analysis, and Excel export.")

    with st.spinner("Fetching live Nike product data..."):
        df = load_line_plan()

    if df.empty:
        st.error("Could not fetch Nike product data. Try again in a moment.")
        return

    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("Total SKUs",       len(df))
    c2.metric("Avg Retail Price", f"${df['Retail Price ($)'].mean():.0f}")
    c3.metric("🆕 New Launches",  int((df["Status"]=="NEW").sum()))
    c4.metric("🔖 On Sale",       int((df["Status"]=="SALE").sum()))
    c5.metric("❌ Out of Stock",  int((df["Status"]=="OUT OF STOCK").sum()))

    st.markdown("---")

    st.sidebar.markdown("### 📋 Line Plan Filters")
    sel_cat    = st.sidebar.selectbox("Category", ["All"]+sorted(df["Category"].dropna().unique().tolist()))
    sel_gender = st.sidebar.selectbox("Gender",   ["All"]+sorted(df["Gender"].dropna().unique().tolist()))
    sel_status = st.sidebar.selectbox("Status",   ["All"]+sorted(df["Status"].dropna().unique().tolist()))
    pmin = float(df["Retail Price ($)"].min() or 0)
    pmax = float(df["Retail Price ($)"].max() or 500)
    price_range = st.sidebar.slider("Price Range ($)", pmin, pmax, (pmin, pmax)) if pmin < pmax else (pmin, pmax)

    filtered = df.copy()
    if sel_cat    != "All": filtered = filtered[filtered["Category"]==sel_cat]
    if sel_gender != "All": filtered = filtered[filtered["Gender"]==sel_gender]
    if sel_status != "All": filtered = filtered[filtered["Status"]==sel_status]
    filtered = filtered[filtered["Retail Price ($)"].between(price_range[0], price_range[1])]

    st.markdown(f"### 🗂️ Line Plan — {len(filtered)} SKUs")
    st.caption("🟢 New · 🔴 Sale · ⬜ Active · ⚫ Out of Stock")

    display_cols = [c for c in ["SKU / Product ID","Product Name","Subtitle",
        "Colorway","Gender","Category","Retail Price ($)","Full Price ($)","Status","In Stock"]
        if c in filtered.columns]

    st.dataframe(filtered[display_cols], column_config={
        "SKU / Product ID":  st.column_config.TextColumn("SKU", width="medium"),
        "Product Name":      st.column_config.TextColumn("Product Name", width="large"),
        "Retail Price ($)":  st.column_config.NumberColumn("Retail $", format="$%.2f"),
        "Full Price ($)":    st.column_config.NumberColumn("Full $",   format="$%.2f"),
        "Status":            st.column_config.TextColumn("Status",     width="small"),
        "In Stock":          st.column_config.TextColumn("Stock",      width="small"),
    }, use_container_width=True, hide_index=True, height=420)

    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("### 💰 Avg Price by Category")
        cp = filtered.groupby("Category")["Retail Price ($)"].mean().sort_values(ascending=False).reset_index()
        cp.columns = ["Category","Avg Price"]
        fig = px.bar(cp, x="Avg Price", y="Category", orientation="h",
                     color="Avg Price", color_continuous_scale=["#60a5fa","#1D9E75"],
                     text=cp["Avg Price"].apply(lambda x: f"${x:.0f}"))
        fig.update_traces(textposition="outside")
        fig.update_layout(height=280, margin=dict(l=0,r=60,t=10,b=0),
                          plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                          coloraxis_showscale=False, yaxis_title="")
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown("### 📊 Status Distribution")
        sc = filtered["Status"].value_counts().reset_index()
        sc.columns = ["Status","Count"]
        fig2 = px.pie(sc, names="Status", values="Count", hole=0.4,
                      color="Status",
                      color_discrete_map={"NEW":"#1D9E75","ACTIVE":"#60a5fa",
                                          "SALE":"#E24B4A","OUT OF STOCK":"#94a3b8"})
        fig2.update_layout(height=280, margin=dict(l=0,r=0,t=10,b=0), paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")
    st.markdown("### 🏷️ Price Band Analysis")
    tmp = filtered.copy()
    tmp["Price Band"] = pd.cut(tmp["Retail Price ($)"],
        bins=[0,75,100,150,200,999], labels=["Under $75","$75–$100","$100–$150","$150–$200","$200+"])
    band = tmp.groupby("Price Band", observed=True).size().reset_index(name="SKUs")
    fig3 = px.bar(band, x="Price Band", y="SKUs", color="SKUs",
                  color_continuous_scale=["#60a5fa","#1D9E75"], text="SKUs")
    fig3.update_traces(textposition="outside")
    fig3.update_layout(height=260, margin=dict(l=0,r=0,t=10,b=0),
                       plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                       coloraxis_showscale=False)
    st.plotly_chart(fig3, use_container_width=True)

    st.markdown("---")
    st.markdown("### 📥 Export Line Plan Report")
    col1, col2 = st.columns([1,3])
    with col1:
        if st.button("Generate Excel Report", type="primary"):
            with st.spinner("Building Excel..."):
                excel_data = export_excel(filtered)
            st.download_button(
                label="📥 Download Nike_Line_Plan.xlsx",
                data=excel_data,
                file_name="Nike_Line_Plan_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    with col2:
        st.info("**Excel sheets:** Line Plan (full SKU list, color-coded) · "
                "Price Analysis · Status Summary · Price Band breakdown")
